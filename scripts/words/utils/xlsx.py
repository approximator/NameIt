#!/usr/bin/env python
# -*- coding: utf-8 -*-
#
# Copyright © 2016 Oleksii Aliakin. All rights reserverd.
# Author: Oleksii Aliakin (alex@nls.la)
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#      http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


class ColIndex:
    word = 1
    part_of_speech = 2
    definition = 3
    example = 4


class Fill:
    red = PatternFill(start_color='FFEE5500', end_color='FFCC0000', fill_type='solid')
    yellow = PatternFill(start_color='FFFFFF00', end_color='FF000000', fill_type='solid')
    green = PatternFill(start_color='FF00BB44', end_color='FF000000', fill_type='solid')


class XlsxFile(object):

    def __init__(self, file_name):
        self.__file_name = file_name

    def write(self, words):

        def fill_row(fill):
            for col in range(1, 10):
                ws_allwords.cell(column=col, row=row_num).fill = fill

        print('Making xlsx...')
        wb = Workbook()
        ws_allwords = wb.active
        ws_allwords.title = 'All words'

        row_num = 1
        ws_allwords.cell(column=ColIndex.word, row=row_num, value='Word')
        ws_allwords.cell(column=ColIndex.part_of_speech, row=row_num, value='Part Of Speech')
        ws_allwords.cell(column=ColIndex.definition, row=row_num, value='Definition')
        ws_allwords.cell(column=ColIndex.example, row=row_num, value='Example')
        ws_allwords.freeze_panes = 'A2'
        ws_allwords.column_dimensions[get_column_letter(ColIndex.word)].width = 15.0
        ws_allwords.column_dimensions[get_column_letter(ColIndex.part_of_speech)].width = 10.0
        ws_allwords.column_dimensions[get_column_letter(ColIndex.definition)].width = 50.0
        ws_allwords.column_dimensions[get_column_letter(ColIndex.example)].width = 50.0

        row_num = 3
        for word in words:
            ws_allwords.cell(column=ColIndex.word, row=row_num, value=word.word)
            ws_allwords.cell(column=ColIndex.part_of_speech, row=row_num, value=word.part_of_speech)
            ws_allwords.cell(column=ColIndex.definition, row=row_num, value=word.definition)
            ws_allwords.cell(column=ColIndex.example, row=row_num, value=word.example)
            if word.part_of_speech and word.definition:
                fill_row(Fill.green)
            if not word.definition:
                fill_row(Fill.yellow)
            if not word.part_of_speech:
                fill_row(Fill.red)

            row_num += 1
        wb.save(filename=self.__file_name)
